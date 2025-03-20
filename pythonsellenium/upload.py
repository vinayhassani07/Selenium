import time

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
# chrome_options.add_argument("headless")
chrome_options.add_argument("--ignore-certificate-errors")
fruit_name = "Apple"
file_path = "E:/vinay/Automation/Seleniumn - 2025/download.xlsx"
book = openpyxl.load_workbook(file_path)
sheet = book.active
Dict = {}
for i in range(2, sheet.max_row + 1):
    if sheet.cell(row=1, column=i).value == "price":
        Dict["col"] = i

for i in range(2, sheet.max_row + 1):
    for j in range(2, sheet.max_column + 1):
        if sheet.cell(row=i, column=j).value == fruit_name:
            Dict["row"] = i

driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(2)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.find_element(By.ID, "downloadButton").click()
sheet.cell(row=Dict["row"], column=Dict["col"]).value = "500"
book.save(file_path)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

WebDriverWait(driver, 10).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast"
                                                                                                    "-body "
                                                                                                    "div:nth-child(2)")))
print(driver.find_element(By.CSS_SELECTOR, ".Toastify__toast"
                                           "-body "
                                           "div:nth-child(2)").text)
